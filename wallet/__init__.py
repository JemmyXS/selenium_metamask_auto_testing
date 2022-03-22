from openpyxl import *
<<<<<<< HEAD
from config import global_config
=======

# filename = '20220312_eth_uniswap_10.xlsx'
>>>>>>> 04ccdd8665eab17bae54d5aa0f127c387123126c


class Excel:

    def __init__(self, file):
        self.file = file
        self.wb = load_workbook(self.file)
        sheets = self.wb.get_sheet_names()
        self.sheet = sheets[0]
        self.ws = self.wb[self.sheet]

    def getColValues(self, column):
        rows = self.ws.max_row
        column_data = []
        for i in range(1, rows + 1):
            cell_value = self.ws.cell(row=i, column=column).value
            column_data.append(cell_value)
        return column_data


<<<<<<< HEAD
def getWallet():
    # 用户助记词路径，以xlsx格式保存，该路径由用户提供,在config.ini中配置
    file = global_config.get('path', 'wallet_path').strip()
    address_list = Excel(file).getColValues(1)
    mnemonic_list = Excel(file).getColValues(3)
    wallet = [address_list, mnemonic_list]
    return wallet


=======
>>>>>>> 04ccdd8665eab17bae54d5aa0f127c387123126c
def getAddress(filename):
    if len(filename) == 0:
        print('未指定地址文件')
        return
    # 用户地址路径，以xlsx格式保存
    file = '/Users/luoye/Downloads/TestNetwork/' + filename
    address_list = Excel(file).getColValues(1)
    return address_list


<<<<<<< HEAD
def getAddressV2():
    wallet = getWallet()
    address_list = wallet[0]
    return address_list


=======
>>>>>>> 04ccdd8665eab17bae54d5aa0f127c387123126c
def getSeedPhrase(filename, address):
    input_address = address
    if len(filename) == 0:
        print('未指定地址文件')
        return
    # 用户助记词路径，以xlsx格式保存，该路径由用户自行修改
    file = '/Users/luoye/Downloads/TestNetwork/' + filename
    address_list = Excel(file).getColValues(1)
    mnemonic_list = Excel(file).getColValues(3)

    for i in range(1, len(address_list)):
        if input_address == address_list[i]:
            num = i
            return mnemonic_list[num]


<<<<<<< HEAD
def getSeedPhraseV2(address):
    input_address = address
    wallet = getWallet()
    address_list = wallet[0]
    mnemonic_list = wallet[1]

    for i in range(1, len(address_list)):
        if input_address == address_list[i]:
            num = i
            return mnemonic_list[num]

=======
def getSeedPhraseV2(filepath):
    # 用户助记词路径，以xlsx格式保存，该路径由用户提供
    # example:'/Users/luoye/Downloads/TestNetwork/20220317_eth_zkSync_muteSwitch_100.xlsx'
    file = filepath
    address_list = Excel(file).getColValues(1)
    mnemonic_list = Excel(file).getColValues(3)
    wallet = [address_list, mnemonic_list]
    return wallet
>>>>>>> 04ccdd8665eab17bae54d5aa0f127c387123126c



